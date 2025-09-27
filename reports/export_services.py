"""
Export services for generating reports in different formats.
Provides services for exporting reports to PDF and Excel formats.
"""

import os
import json
from io import BytesIO
from typing import Dict, Any, Optional
from datetime import datetime
from django.conf import settings
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.utils import timezone

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportService:
    """
    Main service for exporting reports in different formats.
    Coordinates different export formats based on requirements.
    """
    
    def __init__(self):
        self.exporters = {
            'pdf': PDFExporter(),
            'excel': ExcelExporter(),
            'json': JSONExporter(),
        }
    
    def export_report(self, report_data: Dict[str, Any], format: str, title: str = None) -> bytes:
        """
        Export report data in the specified format.
        
        Args:
            report_data: Report data to export
            format: Export format ('pdf', 'excel', 'json')
            title: Optional title for the report
            
        Returns:
            Bytes content of the exported report
            
        Raises:
            ValueError: If format is not supported
            ImportError: If required libraries are not available
        """
        if format not in self.exporters:
            raise ValueError(f"Unsupported export format: {format}")
        
        exporter = self.exporters[format]
        return exporter.export(report_data, title or "Report")
    
    def get_available_formats(self) -> list:
        """
        Get list of available export formats.
        
        Returns:
            List of available format names
        """
        available = ['json']  # JSON is always available
        
        if REPORTLAB_AVAILABLE:
            available.append('pdf')
        
        if OPENPYXL_AVAILABLE:
            available.append('excel')
        
        return available
    
    def get_content_type(self, format: str) -> str:
        """
        Get content type for the specified format.
        
        Args:
            format: Export format
            
        Returns:
            Content type string
        """
        content_types = {
            'pdf': 'application/pdf',
            'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'json': 'application/json',
        }
        return content_types.get(format, 'application/octet-stream')
    
    def get_file_extension(self, format: str) -> str:
        """
        Get file extension for the specified format.
        
        Args:
            format: Export format
            
        Returns:
            File extension string
        """
        extensions = {
            'pdf': '.pdf',
            'excel': '.xlsx',
            'json': '.json',
        }
        return extensions.get(format, '.bin')


class BaseExporter:
    """
    Base class for report exporters.
    Provides common functionality for all export formats.
    """
    
    def export(self, report_data: Dict[str, Any], title: str) -> bytes:
        """
        Export report data.
        
        Args:
            report_data: Report data to export
            title: Report title
            
        Returns:
            Bytes content of the exported report
        """
        raise NotImplementedError("Subclasses must implement export method")
    
    def format_date(self, date_str: str) -> str:
        """
        Format date string for display.
        
        Args:
            date_str: ISO date string
            
        Returns:
            Formatted date string
        """
        try:
            if isinstance(date_str, str):
                date_obj = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                return date_obj.strftime('%d/%m/%Y')
            return str(date_str)
        except (ValueError, AttributeError):
            return str(date_str)
    
    def format_number(self, number: Any) -> str:
        """
        Format number for display.
        
        Args:
            number: Number to format
            
        Returns:
            Formatted number string
        """
        if isinstance(number, (int, float)):
            if isinstance(number, float):
                return f"{number:.2f}"
            return str(number)
        return str(number)


class JSONExporter(BaseExporter):
    """
    Exporter for JSON format.
    Simple JSON serialization of report data.
    """
    
    def export(self, report_data: Dict[str, Any], title: str) -> bytes:
        """
        Export report data as JSON.
        
        Args:
            report_data: Report data to export
            title: Report title
            
        Returns:
            JSON bytes
        """
        export_data = {
            'title': title,
            'exported_at': timezone.now().isoformat(),
            'data': report_data
        }
        
        json_str = json.dumps(export_data, indent=2, ensure_ascii=False, default=str)
        return json_str.encode('utf-8')


class PDFExporter(BaseExporter):
    """
    Exporter for PDF format using ReportLab.
    Creates formatted PDF reports with tables and charts.
    """
    
    def __init__(self):
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab is required for PDF export. Install with: pip install reportlab")
    
    def export(self, report_data: Dict[str, Any], title: str) -> bytes:
        """
        Export report data as PDF.
        
        Args:
            report_data: Report data to export
            title: Report title
            
        Returns:
            PDF bytes
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 12))
        
        # Metadata
        if 'metadata' in report_data:
            metadata = report_data['metadata']
            story.append(Paragraph("Información del Reporte", styles['Heading2']))
            
            metadata_data = [
                ['Generado:', self.format_date(metadata.get('generated_at', ''))],
                ['Período:', f"{self.format_date(metadata.get('date_range', {}).get('start', ''))} - {self.format_date(metadata.get('date_range', {}).get('end', ''))}"],
                ['Total de registros:', str(metadata.get('total_records', 0))]
            ]
            
            metadata_table = Table(metadata_data, colWidths=[2*inch, 3*inch])
            metadata_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(metadata_table)
            story.append(Spacer(1, 20))
        
        # Summary
        if 'summary' in report_data:
            story.append(Paragraph("Resumen", styles['Heading2']))
            summary_table = self._create_summary_table(report_data['summary'])
            story.append(summary_table)
            story.append(Spacer(1, 20))
        
        # Analysis sections
        analysis_sections = [
            ('by_type', 'Análisis por Tipo'),
            ('by_responsible', 'Análisis por Responsable'),
            ('by_genus', 'Análisis por Género'),
            ('by_seed_source', 'Análisis por Fuente de Semillas'),
            ('by_month', 'Análisis Mensual')
        ]
        
        for section_key, section_title in analysis_sections:
            if section_key in report_data and report_data[section_key]:
                story.append(Paragraph(section_title, styles['Heading2']))
                analysis_table = self._create_analysis_table(report_data[section_key], section_key)
                story.append(analysis_table)
                story.append(Spacer(1, 15))
        
        # Success rates
        if 'success_rates' in report_data:
            story.append(Paragraph("Tasas de Éxito", styles['Heading2']))
            success_table = self._create_success_rates_table(report_data['success_rates'])
            story.append(success_table)
            story.append(Spacer(1, 20))
        
        # Records (limited to first 50 for PDF)
        if 'records' in report_data and report_data['records']:
            story.append(Paragraph("Registros Detallados (Primeros 50)", styles['Heading2']))
            records_table = self._create_records_table(report_data['records'][:50])
            story.append(records_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _create_summary_table(self, summary: Dict[str, Any]) -> Table:
        """Create summary table for PDF."""
        data = [
            ['Total de registros:', str(summary.get('total_records', 0))],
            ['Registros únicos:', str(summary.get('unique_records', 0))],
            ['Promedio por día:', str(summary.get('average_per_day', 0))]
        ]
        
        # Add specific fields based on report type
        if 'total_capsules' in summary:
            data.append(['Total de cápsulas:', str(summary.get('total_capsules', 0))])
        if 'total_seedlings' in summary:
            data.append(['Total de plántulas:', str(summary.get('total_seedlings', 0))])
        
        table = Table(data, colWidths=[2.5*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table
    
    def _create_analysis_table(self, analysis_data: list, section_key: str) -> Table:
        """Create analysis table for PDF."""
        if not analysis_data:
            return Table([['No hay datos disponibles']])
        
        # Determine headers based on section
        if section_key == 'by_type':
            headers = ['Tipo', 'Cantidad', 'Porcentaje', 'Promedio']
            data = [[item.get('type', ''), str(item.get('count', 0)), 
                    f"{item.get('percentage', 0)}%", 
                    str(item.get('avg_capsules', item.get('avg_seedlings', 0)))] 
                   for item in analysis_data]
        elif section_key == 'by_responsible':
            headers = ['Usuario', 'Nombre', 'Cantidad', 'Promedio']
            data = [[item.get('username', ''), item.get('full_name', ''), 
                    str(item.get('count', 0)), 
                    str(item.get('avg_capsules', item.get('avg_seedlings', 0)))] 
                   for item in analysis_data]
        elif section_key == 'by_genus':
            headers = ['Género', 'Cantidad', 'Promedio']
            data = [[item.get('genus', ''), str(item.get('count', 0)), 
                    str(item.get('avg_capsules', item.get('avg_seedlings', 0)))] 
                   for item in analysis_data]
        elif section_key == 'by_seed_source':
            headers = ['Fuente', 'Cantidad', 'Promedio']
            data = [[item.get('source', ''), str(item.get('count', 0)), 
                    str(item.get('avg_seedlings', 0))] 
                   for item in analysis_data]
        elif section_key == 'by_month':
            headers = ['Mes', 'Cantidad', 'Promedio']
            data = [[item.get('month', ''), str(item.get('count', 0)), 
                    str(item.get('avg_capsules', item.get('avg_seedlings', 0)))] 
                   for item in analysis_data]
        else:
            headers = ['Campo', 'Valor']
            data = [[str(k), str(v)] for item in analysis_data for k, v in item.items()]
        
        table_data = [headers] + data
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table
    
    def _create_success_rates_table(self, success_rates: Dict[str, Any]) -> Table:
        """Create success rates table for PDF."""
        data = [
            ['Total de registros:', str(success_rates.get('total_records', 0))],
            ['Registros exitosos:', str(success_rates.get('successful_records', 0))],
            ['Tasa de éxito:', f"{success_rates.get('success_rate', 0)}%"]
        ]
        
        table = Table(data, colWidths=[2.5*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table
    
    def _create_records_table(self, records: list) -> Table:
        """Create records table for PDF."""
        if not records:
            return Table([['No hay registros disponibles']])
        
        # Get headers from first record
        first_record = records[0]
        headers = list(first_record.keys())
        
        # Limit headers for PDF readability
        important_headers = ['date', 'responsible', 'type', 'observations']
        filtered_headers = [h for h in headers if h in important_headers][:4]
        
        data = []
        for record in records:
            row = [str(record.get(header, ''))[:30] for header in filtered_headers]  # Truncate long text
            data.append(row)
        
        table_data = [filtered_headers] + data
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table


class ExcelExporter(BaseExporter):
    """
    Exporter for Excel format using openpyxl.
    Creates formatted Excel workbooks with multiple sheets.
    """
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    def export(self, report_data: Dict[str, Any], title: str) -> bytes:
        """
        Export report data as Excel.
        
        Args:
            report_data: Report data to export
            title: Report title
            
        Returns:
            Excel bytes
        """
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create summary sheet
        self._create_summary_sheet(workbook, report_data, title)
        
        # Create analysis sheets
        analysis_sections = [
            ('by_type', 'Análisis por Tipo'),
            ('by_responsible', 'Por Responsable'),
            ('by_genus', 'Por Género'),
            ('by_seed_source', 'Por Fuente'),
            ('by_month', 'Análisis Mensual')
        ]
        
        for section_key, sheet_name in analysis_sections:
            if section_key in report_data and report_data[section_key]:
                self._create_analysis_sheet(workbook, report_data[section_key], sheet_name, section_key)
        
        # Create records sheet
        if 'records' in report_data and report_data['records']:
            self._create_records_sheet(workbook, report_data['records'])
        
        # Save to bytes
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _create_summary_sheet(self, workbook, report_data: Dict[str, Any], title: str):
        """Create summary sheet in Excel workbook."""
        sheet = workbook.create_sheet("Resumen")
        
        # Title
        sheet['A1'] = title
        sheet['A1'].font = Font(size=16, bold=True)
        sheet.merge_cells('A1:D1')
        
        # Metadata
        row = 3
        if 'metadata' in report_data:
            metadata = report_data['metadata']
            sheet[f'A{row}'] = "Información del Reporte"
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            sheet[f'A{row}'] = "Generado:"
            sheet[f'B{row}'] = self.format_date(metadata.get('generated_at', ''))
            row += 1
            
            date_range = metadata.get('date_range', {})
            sheet[f'A{row}'] = "Período:"
            sheet[f'B{row}'] = f"{self.format_date(date_range.get('start', ''))} - {self.format_date(date_range.get('end', ''))}"
            row += 1
            
            sheet[f'A{row}'] = "Total de registros:"
            sheet[f'B{row}'] = metadata.get('total_records', 0)
            row += 2
        
        # Summary
        if 'summary' in report_data:
            summary = report_data['summary']
            sheet[f'A{row}'] = "Resumen Estadístico"
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for key, value in summary.items():
                sheet[f'A{row}'] = key.replace('_', ' ').title()
                sheet[f'B{row}'] = value
                row += 1
        
        # Auto-adjust column widths
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = None
            for cell in column_cells:
                # Skip merged cells
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_analysis_sheet(self, workbook, analysis_data: list, sheet_name: str, section_key: str):
        """Create analysis sheet in Excel workbook."""
        sheet = workbook.create_sheet(sheet_name)
        
        if not analysis_data:
            sheet['A1'] = "No hay datos disponibles"
            return
        
        # Headers
        headers = self._get_analysis_headers(section_key)
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, item in enumerate(analysis_data, 2):
            values = self._get_analysis_values(item, section_key)
            for col, value in enumerate(values, 1):
                sheet.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = None
            for cell in column_cells:
                # Skip merged cells
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 30)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_records_sheet(self, workbook, records: list):
        """Create records sheet in Excel workbook."""
        sheet = workbook.create_sheet("Registros Detallados")
        
        if not records:
            sheet['A1'] = "No hay registros disponibles"
            return
        
        # Headers
        headers = list(records[0].keys())
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header.replace('_', ' ').title())
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, record in enumerate(records, 2):
            for col, header in enumerate(headers, 1):
                value = record.get(header, '')
                if isinstance(value, str) and len(value) > 100:
                    value = value[:100] + "..."  # Truncate long text
                sheet.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = None
            for cell in column_cells:
                # Skip merged cells
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 40)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _get_analysis_headers(self, section_key: str) -> list:
        """Get headers for analysis section."""
        headers_map = {
            'by_type': ['Tipo', 'Cantidad', 'Porcentaje', 'Promedio'],
            'by_responsible': ['Usuario', 'Nombre Completo', 'Cantidad', 'Promedio'],
            'by_genus': ['Género', 'Cantidad', 'Promedio'],
            'by_seed_source': ['Fuente', 'Cantidad', 'Promedio'],
            'by_month': ['Mes', 'Cantidad', 'Promedio']
        }
        return headers_map.get(section_key, ['Campo', 'Valor'])
    
    def _get_analysis_values(self, item: Dict[str, Any], section_key: str) -> list:
        """Get values for analysis item."""
        if section_key == 'by_type':
            return [
                item.get('type', ''),
                item.get('count', 0),
                f"{item.get('percentage', 0)}%",
                item.get('avg_capsules', item.get('avg_seedlings', 0))
            ]
        elif section_key == 'by_responsible':
            return [
                item.get('username', ''),
                item.get('full_name', ''),
                item.get('count', 0),
                item.get('avg_capsules', item.get('avg_seedlings', 0))
            ]
        elif section_key == 'by_genus':
            return [
                item.get('genus', ''),
                item.get('count', 0),
                item.get('avg_capsules', item.get('avg_seedlings', 0))
            ]
        elif section_key == 'by_seed_source':
            return [
                item.get('source', ''),
                item.get('count', 0),
                item.get('avg_seedlings', 0)
            ]
        elif section_key == 'by_month':
            return [
                item.get('month', ''),
                item.get('count', 0),
                item.get('avg_capsules', item.get('avg_seedlings', 0))
            ]
        else:
            return [str(k) for k in item.values()]